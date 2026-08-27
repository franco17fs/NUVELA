#!/usr/bin/env python3
"""
Arma el .xlsx del sistema de cashflow desde el modelo exportado de los .gs.

Se sube a Google Drive y Drive lo convierte a Google Sheets, así queda la
planilla lista sin cargar nada a mano. El Apps Script se pega aparte cuando
haga falta el motor: no hay API para adjuntarlo.

Uso:
    node cashflow/build/exportar-modelo.js > modelo.json
    python3 cashflow/build/construir_xlsx.py modelo.json salida.xlsx
"""
import datetime as dt
import json
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Paleta de marca NUVELA (manual de marca).
AZUL_ENTRADA = "232B52"
GRIS_GENERADA = "5A6070"
CELESTE_TAB = "92ACD8"

FORMATO = {'"$"#,##0': '"$"#,##0', "dd/mm/yyyy": "DD/MM/YYYY"}


def es_fecha(valor):
    return isinstance(valor, str) and len(valor) == 10 and valor[4] == "-" and valor[7] == "-"


def construir(modelo, destino):
    wb = Workbook()
    wb.remove(wb.active)

    for clave in modelo["orden"]:
        definicion = modelo["esquema"][clave]
        columnas = definicion["columnas"]
        hoja = wb.create_sheet(definicion["nombre"])
        generada = definicion["generada"]

        hoja.sheet_properties.tabColor = GRIS_GENERADA if generada else CELESTE_TAB

        relleno = PatternFill("solid", fgColor=GRIS_GENERADA if generada else AZUL_ENTRADA)
        for i, col in enumerate(columnas, start=1):
            celda = hoja.cell(row=1, column=i, value=col["titulo"])
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = relleno
            celda.alignment = Alignment(vertical="center", wrap_text=True)
            if col.get("nota"):
                celda.comment = Comment(col["nota"], "NUVELA Cashflow", height=140, width=380)
            hoja.column_dimensions[get_column_letter(i)].width = max(col["ancho"] / 7.0, 9)

        hoja.freeze_panes = "A2"
        hoja.row_dimensions[1].height = 34

        filas = modelo["semillas"].get(clave, [])
        for f, fila in enumerate(filas, start=2):
            for c, valor in enumerate(fila, start=1):
                if es_fecha(valor):
                    valor = dt.date.fromisoformat(valor)
                celda = hoja.cell(row=f, column=c, value=valor)
                formato = columnas[c - 1].get("formato")
                if formato:
                    celda.number_format = FORMATO[formato]
                if columnas[c - 1]["titulo"] in ("Notas", "Consecuencia_Atraso", "Qué significa",
                                                 "Forma_Cobro", "Qué significa "):
                    celda.alignment = Alignment(vertical="top", wrap_text=True)

        # Las validaciones se aplican a un rango amplio para que las filas
        # nuevas también hereden el desplegable.
        ultima = max(len(filas) + 200, 300)
        for i, col in enumerate(columnas, start=1):
            if not col.get("lista"):
                continue
            opciones = ",".join(str(v) for v in col["lista"])
            dv = DataValidation(type="list", formula1=f'"{opciones}"', allow_blank=True,
                                showDropDown=False, showErrorMessage=False)
            hoja.add_data_validation(dv)
            letra = get_column_letter(i)
            dv.add(f"{letra}2:{letra}{ultima}")

        if generada and not filas:
            aviso = hoja.cell(row=2, column=1,
                              value="Esta hoja la escribe el sistema. Se completa en la etapa siguiente.")
            aviso.font = Font(italic=True, color="8A8F9A")

    wb.save(destino)
    return wb


if __name__ == "__main__":
    modelo = json.load(open(sys.argv[1], encoding="utf-8"))
    wb = construir(modelo, sys.argv[2])
    print(f"{sys.argv[2]}: {len(wb.sheetnames)} hojas -> {', '.join(wb.sheetnames)}")
