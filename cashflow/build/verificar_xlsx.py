#!/usr/bin/env python3
"""
Verifica el .xlsx generado antes de subirlo a Drive.

Chequea lo que rompería la planilla en silencio: hojas faltantes, formatos que
no llegan, semanas mal generadas y —sobre todo— las tres trampas de doble
conteo del modelo.

Uso:  python3 cashflow/build/verificar_xlsx.py modelo.json salida.xlsx
"""
import datetime as dt
import json
import sys

from openpyxl import load_workbook

fallas = []


def chk(cond, msg):
    print(("  OK   " if cond else "FALLA  ") + msg)
    if not cond:
        fallas.append(msg)


def verificar(modelo, ruta):
    wb = load_workbook(ruta)
    esperadas = [modelo["esquema"][c]["nombre"] for c in modelo["orden"]]
    chk(wb.sheetnames == esperadas, f"hojas en orden: {wb.sheetnames}")

    for clave in modelo["orden"]:
        definicion = modelo["esquema"][clave]
        hoja = wb[definicion["nombre"]]
        columnas = definicion["columnas"]
        titulos = [c.value for c in hoja[1][: len(columnas)]]
        chk(titulos == [c["titulo"] for c in columnas], f"{definicion['nombre']}: cabecera completa")
        chk(hoja.freeze_panes == "A2", f"{definicion['nombre']}: cabecera congelada")
        filas = modelo["semillas"].get(clave, [])
        if filas:
            chk(hoja.max_row == len(filas) + 1,
                f"{definicion['nombre']}: {len(filas)} filas de datos")

    # --- Las tres trampas de doble conteo -----------------------------------
    obl = {r[0]: r for r in wb["Obligaciones"].iter_rows(min_row=2, values_only=True)}
    chk(obl["OBL-002"][7] < 1_000_000,
        "OBL-002 carga el saldo adeudado, no el total facturado del ciclo (~$4.400.000)")
    chk(obl["OBL-012"][1] == "NO",
        "OBL-012 (Ads) desactivada: ya viene dentro de la factura de ML")
    activas = [r for r in obl.values() if r[1] == "SI"]
    chk(not any("auto" in str(r[2]).lower() or "galicia" in str(r[2]).lower() for r in activas),
        "la cuota del auto no figura como obligación: se paga del retiro")

    chk(all(str(r[11]).strip() for r in activas),
        "toda obligación activa explica su consecuencia de atraso")
    chk(all(1 <= int(r[5]) <= 5 for r in activas), "criticidades entre 1 y 5")

    # --- Semanas -------------------------------------------------------------
    ventas = wb["Ventas"]
    desde, hasta = ventas["B2"].value, ventas[f"C{ventas.max_row}"].value
    chk(isinstance(desde, dt.datetime) and desde.weekday() == 0, f"semana 1 arranca lunes ({desde:%d/%m/%Y})")
    chk(isinstance(hasta, dt.datetime) and hasta.weekday() == 6, f"última semana cierra domingo ({hasta:%d/%m/%Y})")
    chk((hasta - desde).days + 1 == 91, "las 13 semanas cubren 91 días sin huecos")

    # --- Formatos y ayudas ---------------------------------------------------
    chk(ventas["D2"].number_format.startswith('"$"'), "los montos salen con formato moneda")
    chk(ventas["B2"].number_format == "DD/MM/YYYY", "las fechas salen dd/mm/aaaa")

    notas_esperadas = sum(1 for c in modelo["orden"]
                          for col in modelo["esquema"][c]["columnas"] if col.get("nota"))
    notas = sum(1 for h in wb for fila in h.iter_rows(max_row=1) for cel in fila if cel.comment)
    chk(notas == notas_esperadas, f"notas explicativas en cabeceras: {notas}/{notas_esperadas}")

    desplegables = sum(len(wb[modelo["esquema"][c]["nombre"]].data_validations.dataValidation)
                       for c in modelo["orden"])
    listas = sum(1 for c in modelo["orden"]
                 for col in modelo["esquema"][c]["columnas"] if col.get("lista"))
    chk(desplegables == listas, f"desplegables: {desplegables}/{listas}")

    for clave in modelo["orden"]:
        if modelo["esquema"][clave]["generada"]:
            hoja = wb[modelo["esquema"][clave]["nombre"]]
            chk("sistema" in str(hoja["A2"].value),
                f"{hoja.title}: avisa que la escribe el sistema")

    # --- Config --------------------------------------------------------------
    cfg = {r[0]: r for r in wb["Config"].iter_rows(min_row=2, values_only=True)}
    chk(all(r[3] and len(r[3]) > 20 for r in cfg.values()), "todo parámetro explica qué significa")
    pendientes = [k for k, r in cfg.items() if r[4] == "CONFIRMAR"]
    chk(len(pendientes) > 0, f"quedan marcados para confirmar: {', '.join(pendientes)}")


if __name__ == "__main__":
    verificar(json.load(open(sys.argv[1], encoding="utf-8")), sys.argv[2])
    print("\n" + ("TODO OK" if not fallas else f"{len(fallas)} FALLAS"))
    sys.exit(0 if not fallas else 1)
